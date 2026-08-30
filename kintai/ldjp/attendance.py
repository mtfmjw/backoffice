from pathlib import Path
from django.contrib import admin
from common.utils import convert2localtime
from openpyxl.worksheet.worksheet import Worksheet as WorkSheet
from datetime import date
from common.models import Member
from kintai.models.monthly_attendance import MonthlyAttendance
from .const import ATTENDANCE_SHEET


def get_organization_code(member: Member) -> str:
    """部署コード"""
    return member.organization.code

def get_member_no(member: Member) -> str:
    """社員番号"""
    return member.user.username

def get_member_name(member: Member) -> str:
    """氏名"""
    return f"{member.user.last_name}{member.user.first_name}"

def get_attendance_sheet_file_name(month: date, member: Member) -> str:
    month_str = month.strftime("%y%m")
    code = get_organization_code(member)
    no = get_member_no(member)
    name = get_member_name(member)
    return f"{ATTENDANCE_SHEET[:8]}_{month_str}_{code}_{no}_{name}.xlsx"

def write_attendance_sheet(ws: WorkSheet, model_admin: admin.ModelAdmin, monthly_attendance: MonthlyAttendance) -> Path:
    """勤怠表を生成する"""
    month = monthly_attendance.month
    member = monthly_attendance.member

    # 2. Populate cells with object data
    ws['AE3'] = get_organization_code(member) # 部署
    ws['AE4'] = get_member_no(member) # 社員番号
    ws['AF4'] = get_member_name(member) # 氏名
    ws['AO3'] = month.strftime("%Y") # 対象年
    ws['AO4'] = month.strftime("%m") # 対象月

    first_row = 11  # Assuming the first row of data starts at row 12
    for daily in monthly_attendance.daily_attendances.all():
        row = daily.day.day + first_row
        ws[f'E{row}'] = daily.work_pattern.no if daily.work_pattern else '' # 就業
        ws[f'F{row}'] = convert2localtime(daily.clock_in_time).strftime("%H:%M") if daily.clock_in_time else '' # 出勤
        ws[f'G{row}'] = convert2localtime(daily.clock_out_time).strftime("%H:%M") if daily.clock_out_time else '' # 退勤
        if daily.date_status and daily.date_status >= 100:  # 特別休暇、育児休業など
            ws[f'I{row}'] = daily.date_status.label  # 特記事項
        if daily.date_status and daily.date_status < 100:  # 出勤、欠勤、有給休暇など
            ws[f'J{row}'] = daily.date_status # 項目番号
        ws[f'AV{row}'] = '1' if not daily.has_break1 and daily.is_present() else '' # 休憩2
        ws[f'AW{row}'] = '1' if not daily.has_break2 and daily.is_present() else '' # 休憩3
        ws[f'AX{row}'] = '1' if not daily.has_break3 and daily.is_present() else '' # 休憩4
        ws[f'AY{row}'] = '1' if not daily.has_break4 and daily.is_present() else '' # 休憩5
        ws[f'AZ{row}'] = '1' if not daily.has_break5 and daily.is_present() else '' # 休憩6